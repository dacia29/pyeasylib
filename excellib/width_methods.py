import os
import openpyxl
import re
from win32com.client import Dispatch
from datetime import datetime

logfile = r"\\Devluna\d$\CODES_SJ\lunaflow\personal_workspace\logs\python_test.txt"

class ColumnsWidthAdjuster:
    '''
    A class to handle Excel file operations using win32com.client.

    Attributes:
        excelfp (str): The file path to the Excel file.
        
    REFERENCES:
    # https://stackoverflow.com/questions/62505403/using-python-win32com-to-get-list-of-excel-worksheets#:~:text=To%20get%20the%20name%20of%20each%20sheet%2C%20you,each%20sheet%2C%20you%20must%20use%20the%20.Name%20method
    # https://stackoverflow.com/questions/13197574/openpyxl-adjust-column-width-size
    # CHATGPT for docstring
    
    #CHANGE LOGS
    #20230322 - Initialised.
    '''
    
    def __init__(self, excelfp):
        '''
        Initializes an instance of the ExcelHandler class.

        Parameters:
            excelfp (str): The file path to the Excel file.
        '''
        
        self.excelfp = excelfp

    def main(self, sheetnames = None, method = None):

        sheetnames = sheetnames

        if method == None:

            f = open(logfile,'a')
            f.write(f"[{str(datetime.now())}] WIN32COM METHOD STARTING...\n")
            f.close()

            self.autofit_win32com(sheetnames)

        elif method == 'openpyxl':

            self.autofit_openpyxl(sheetnames)

        else:

            msg = f'Method "{method}" not recognised.'
            f = open(logfile,'a')
            f.write(f"[{str(datetime.now())}] {msg}\n")
            f.close()

    def is_cell_in_merged_range(self, ws, cell):
        # Iterate over all merged cell ranges in the worksheet
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                return True
        return False
    
    def is_number(self, s):
            try:
                float(s)
                return True
            except ValueError:
                return False
            
    def get_merged_ranges(self, sheet):
            merged_ranges = set()
            for merged_range in sheet.merged_cells.ranges:
                for cell in merged_range.cells:
                    merged_ranges.add(cell)
            return merged_ranges
    
    def autofit_win32com(self, sheetnames=None):
        '''
        Performs autofit operations on the specified sheets of the Excel file.

        Parameters:
            sheetnames (list or bool, optional): A list of sheet names to autofit. 
                If True, autofits all sheets. If None or False, no operation is performed. 
                Defaults to None.
        '''
        if not sheetnames:
            print("No sheet names provided. No need.")
            return
        
        try:
            f = open(logfile,'a')
            f.write(f"1\n")
            f.close()
            excel = Dispatch('Excel.Application')
            excel.Visible = False  # Set Excel to run in the background
            excel.DisplayAlerts = False
            f = open(logfile,'a')
            f.write(f"2\n")
            f.close()
            # win 32 only works with absolute path
            self.excelfp_abs = os.path.abspath(self.excelfp)
            
            print(f"Opening Excel file: {self.excelfp_abs}...")

            f = open(logfile,'a')
            f.write(f"3\n")
            f.close()

            try:
                wb = excel.Workbooks.Open(self.excelfp_abs)

            except Exception as e:
                # f = open(logfile,'a')
                # f.write(f"{e}\n")
                # f.close()
                print(e)

            f = open(logfile,'a')
            f.write(f"4\n")
            f.close()
            
            if isinstance(sheetnames, list):
                sn_list = sheetnames
            elif sheetnames is True:
                sn_list = [ws.Name for ws in wb.Sheets]
            else:
                raise ValueError(f"Invalid type for sheetnames: {type(sheetnames)}")
            
            f = open(logfile,'a')
            f.write(f"5\n")
            f.close()

            for sheetname in sn_list:
                try:
                    ws = excel.Worksheets(sheetname)
                    ws.Activate()
                    ws.Columns.AutoFit()
                    ws.Rows.AutoFit()
                    print(f"AutoFit done for sheet '{sheetname}'.")
                except Exception as e:
                    print(f"Error while processing sheet '{sheetname}': {e}")
            
            f = open(logfile,'a')
            f.write(f"6\n")
            f.close()

            # wb.Save()

            # create new filepath to save it in
            #   temp solution until we figure out why the file cannot be overwritten
            dirname, filename = os.path.split(self.excelfp_abs)
            # Split the file name into the name and extension
            name, ext = os.path.splitext(filename)
            # Create the new file name by adding the suffix
            new_filename = f"{name}_formatted{ext}"
            # Combine the directory and new file name to get the new file path
            new_filepath = os.path.join(dirname, new_filename)

            f = open(logfile,'a')
            f.write(f"7\n")
            f.close()

            wb.SaveAs(new_filepath)
            wb.Close()
            excel.Quit()
            # print(f"Excel file saved and closed successfully at: {self.excelfp_abs}.")
            # print(f"Excel file saved and closed successfully at: {new_filepath}.")

            # assert False
            
        except Exception as e:
            # f = open(logfile,'a')
            # f.write(f"{e}\n")
            # f.close()
            print(f"An error occurred: {e}")
            f = open(logfile,'a')
            f.write(f"An error has occurred. Please approach DS team.\n")
            f.close()
            # wb.Close()
            # excel.Quit()
            # print("Closed the file successfully")


    def autofit_openpyxl(self, sheetnames = None):
        # wb = openpyxl.load_workbook(self.excelfp)        
        
        # if isinstance(sheetnames, list):
        #         sn_list = sheetnames
        # elif sheetnames is True:
        #     sn_list = wb.sheetnames
        # else:
        #     raise ValueError(f"Invalid type for sheetnames: {type(sheetnames)}")

        # for sheetname in sn_list:
        #     ws = wb[sheetname]

        #     max_column = ws.max_column
        #     max_row = ws.max_row

        #     #assert False
        #     for col_index in range(1, max_column+1):
                
        #         col_alpha = openpyxl.utils.get_column_letter(col_index)

        #         max_width = 0
        #         for row_index in range(1, max_row+1):
        #             #value - ignore value if value reside in merged cell
        #             value = ws.cell(row_index, col_index).value if not self.is_cell_in_merged_range(ws, ws[col_alpha+str(row_index)]) else ""
        #             value2 = "" if value is None else value
        #             #value3 - removing HH:MM:SS from datetime string
        #             value3 = re.sub(r"\s*[\d]{2}\:[\d]{2}\:[\d]{2}$","",str(value)) if re.match(r".*\s*[\d]{2}\:[\d]{2}\:[\d]{2}$",str(value)) else value2
        #             #value4 - removing decimals from numeric string (if any)
        #             value4 = round(float(value),2) if self.is_number(str(value)) else value3
        #             #value5 - checking for formula and ignore if found
        #             value5 = re.sub(r"^\=.*","",str(value)) if re.match(r"^\=.*",str(value)) else value4

        #             width = len(str(value5))

                    
        #             # update width
        #             max_width = max(max_width, width)
                    
                
                                
        #             print (col_alpha, col_index, row_index, value, value2, "|", width, max_width)
                    
        #         # adjust
        #         adjusted_width = (max_width + 2) * 1.2
        #         if adjusted_width < 50:
        #             ws.column_dimensions[col_alpha].width = adjusted_width
        #         else:
        #             ws.column_dimensions[col_alpha].width = 50
                
        #         print (col_alpha, adjusted_width)
                
        #         print ("-" * 50)
            
    
        # wb.save(self.excelfp)

        # Load the workbook once at the start
        wb = openpyxl.load_workbook(self.excelfp)

        # Determine the list of sheet names based on input
        if isinstance(sheetnames, list):
            sn_list = sheetnames
        elif sheetnames is True:
            sn_list = wb.sheetnames
        else:
            raise ValueError(f"Invalid type for sheetnames: {type(sheetnames)}")

        # Precompiled regex patterns for performance
        time_pattern = re.compile(r"\s*[\d]{2}:[\d]{2}:[\d]{2}$")
        formula_pattern = re.compile(r"^\=.*")

        for sheetname in sn_list:
            ws = wb[sheetname]
            max_column = ws.max_column
            max_row = ws.max_row

            # Precompute merged ranges once per sheet
            merged_ranges = self.get_merged_ranges(ws)

            for col_index in range(1, max_column + 1):
                col_alpha = openpyxl.utils.get_column_letter(col_index)
                max_width = 0
                column_widths = []

                for row_index in range(1, max_row + 1):
                    cell_position = (row_index, col_index)
                    cell = ws.cell(row_index, col_index)

                    # Skip merged cells by checking the precomputed set
                    if cell_position in merged_ranges:
                        value = ""
                    else:
                        value = cell.value

                    # Value processing
                    if value is None:
                        value_processed = ""
                    else:
                        value_str = str(value)
                        # Remove time from datetime string
                        value_processed = time_pattern.sub("", value_str) if time_pattern.search(value_str) else value_str
                        # Remove decimals from numeric string
                        if self.is_number(value_processed):
                            value_processed = round(float(value_processed), 2)
                            value_processed = str(value_processed)
                        # Remove formula if present
                        if formula_pattern.match(value_processed):
                            value_processed = formula_pattern.sub("", value_processed)

                    column_widths.append(len(value_processed))  # Add the length to list

                # Get max width for column after iterating all rows
                max_width = max(column_widths, default=0)

                # Adjust the width of the column
                adjusted_width = min((max_width + 2) * 1.2, 50)
                ws.column_dimensions[col_alpha].width = adjusted_width

        # Save the workbook
        wb.save(self.excelfp)

if __name__ == "__main__":
    
    
    # Testing adjustment for column width
    if False:
        
        fp = r"./test/file3 - adjust width.xlsx"
        copied_fp = r"./test/file3 - adjust width output.xlsx"

        # Make a copy of the file
        wb = openpyxl.open(fp)
        wb.save(copied_fp)

        # Adjust width on the copied file
        self = ColumnsWidthAdjuster(copied_fp)
        self.main(["Data2"])
    
        assert False, "End of script."
    
    # Non win32 method - KIV only
    if True:
        
        fp = r"./test/file3 - adjust width.xlsx"
        copied_fp = r"./test/file3 - adjust width output.xlsx"

        # Make a copy of the file
        wb = openpyxl.open(fp)
        wb.save(copied_fp)
        
        self = ColumnsWidthAdjuster(copied_fp)
        self.main(["Data2"], method = "openpyxl")